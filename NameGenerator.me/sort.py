from openpyxl import load_workbook, Workbook
import os

# أسماء الملفات الأربعة
files = [
    'UNIQUEULTIMATEV2.xlsx',
    'UNIQUEULTIMATEV3.xlsx',
    'UNIQUEULTIMATEV366.xlsx',
]

# لون الأحمر (عدّل إذا كان لونك مختلف)
RED_COLORS = {'#bf0041'}

# إنشاء ملف جديد
new_wb = Workbook()
new_ws = new_wb.active
new_ws.title = 'Red Cells'

new_ws['A1'] = 'Red Highlighted Cells'

row_out = 2

for filename in files:
    if not os.path.exists(filename):
        print(f'{filename} not found')
        continue

    wb = load_workbook(filename)
    ws = wb.active

    for row in ws.iter_rows():
        for cell in row:
            if cell.value is None:
                continue

            fill = cell.fill
            if fill and fill.fill_type == 'solid':
                color = fill.fgColor.rgb
                if color in RED_COLORS:
                    new_ws.cell(row=row_out, column=1, value=cell.value)
                    row_out += 1

# حفظ الملف
new_wb.save('sort12000.xlsx')
print('Done. Saved all red highlighted cells into sort.xlsx')