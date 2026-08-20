import random

# =========================
# إعدادات
# =========================

SEED = None          # ضع رقم مثل 42 إذا أردت نتائج ثابتة
COUNT = 3000         # عدد الأسماء
LENGTHS = [3, 4, 5, 6]  # أطوال الأسماء

if SEED is not None:
    random.seed(SEED)

# =========================
# الرموز
# =========================

symbols = [
    '⏆', '⌱', '⏚', '⌰', '⍤', '⌬', '⍣', '⌔', '⌕', '⍬',
    '⌕', '⌀', '⌁', '⌭', '⍝', '⫘', '⨱', '⩕', '⩙', '⩖',
    '⨰', '⨔', '⨰', '⨱', '⋰', '⋱', '⊶', '⊷', '⋄', '⋆',
    '⋇', '⥈', '⥉', '⥉', '⥾', '⥿', '⥺', '⥉', '⥈', '⅊',
    '⤙', '⤚', '‡', '‽', '⁂', '⁜'
]

# =========================
# الحروف مع أوزان
# =========================

vowels = {
    'a': 10, 'e': 10, 'i': 8, 'o': 8, 'u': 4,
    'y': 3, 'æ': 2, 'ø': 2, 'ê': 1, 'ï': 1
}

consonants = {
    'b': 3, 'c': 4, 'd': 5, 'f': 3, 'g': 4,
    'h': 4, 'j': 1, 'k': 6, 'l': 8, 'm': 6,
    'n': 8, 'p': 4, 'q': 1, 'r': 9, 's': 8,
    't': 8, 'v': 4, 'w': 2, 'x': 1, 'z': 2,
    'ç': 1
}

# =========================
# الأنماط
# =========================

patterns = {
    3: [('CVC', 6), ('CVV', 2), ('VCV', 2)],
    4: [('CVCV', 5), ('CVCC', 3), ('CCVC', 2)],
    5: [('CVCVC', 5), ('CVCCV', 3), ('CCVCV', 2)],
    6: [('CVCVCV', 5), ('CVCCVC', 3), ('CCVCVC', 2)]
}

good_bigrams = {
    'th','sh','ch','ar','er','en','or','an','el','ri','ra',
    'li','ka','va','ir','al','on','is','os','un','yn','ae'
}

bad_bigrams = {
    'qx','xq','zj','jq','ww','qq','xx','zz','vv'
}

# =========================
# أدوات
# =========================

def weighted_choice(weight_dict):
    items = list(weight_dict.keys())
    weights = list(weight_dict.values())
    return random.choices(items, weights=weights, k=1)[0]

def choose_pattern(length):
    p = patterns[length]
    names = [x[0] for x in p]
    weights = [x[1] for x in p]
    return random.choices(names, weights=weights, k=1)[0]

def generate_base(length):
    pattern = choose_pattern(length)
    result = []

    for ch in pattern:
        if ch == 'C':
            result.append(weighted_choice(consonants))
        else:
            result.append(weighted_choice(vowels))

    return ''.join(result)

def pronounceable(name):
    vowel_set = set(vowels.keys())

    if not any(c in vowel_set for c in name):
        return False

    cons = 0
    vows = 0

    for c in name:
        if c in vowel_set:
            vows += 1
            cons = 0
        else:
            cons += 1
            vows = 0

        if cons > 2 or vows > 2:
            return False

    for bg in bad_bigrams:
        if bg in name:
            return False

    return True

def score_name(name):
    score = 0

    if name[0] in 'klmnrstva':
        score += 2

    for bg in good_bigrams:
        if bg in name:
            score += 3

    if len(set(name)) == len(name):
        score += 2

    if name[-1] in 'aeinor':
        score += 2

    if 'q' in name and 'qu' not in name:
        score -= 5

    return score

def stylize(name):
    name = name.capitalize()

    # احتمال 6% لإضافة رمز
    if random.random() < 0.06:
        sym = random.choice(symbols)

        mode = random.choice(['prefix','suffix','surround'])

        if mode == 'prefix':
            name = sym + name
        elif mode == 'suffix':
            name = name + sym
        else:
            name = sym + name + random.choice(symbols)

    return name

# =========================
# التوليد
# =========================

generated = set()
results = []

attempts = 0

while len(results) < COUNT and attempts < COUNT * 30:
    attempts += 1

    length = random.choice(LENGTHS)
    base = generate_base(length)

    if not pronounceable(base):
        continue

    if score_name(base) < 4:
        continue

    if base in generated:
        continue

    generated.add(base)
    results.append(stylize(base))

# =========================
# حفظ الملف
# =========================
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

# إنشاء الملف
wb = Workbook()
ws = wb.active
ws.title = 'Generated Names'

# العناوين
headers = ['ID', 'Name', 'Length', 'Pattern', 'Score', 'Rarity', 'Symbol']
for col, header in enumerate(headers, 1):
    cell = ws.cell(row=1, column=col, value=header)
    cell.font = Font(bold=True, color='FFFFFF')
    cell.fill = PatternFill('solid', fgColor='1F4E78')
    cell.alignment = Alignment(horizontal='center')

# ألوان الندرة
rarity_colors = {
    'Common': 'D9EAD3',
    'Rare': '9FC5E8',
    'Epic': 'C27BA0',
    'Legendary': 'FFD966'
}

# تحديد الندرة من السكور
def rarity_from_score(score, has_symbol):
    if score >= 12 and has_symbol:
        return 'Legendary'
    elif score >= 10:
        return 'Epic'
    elif score >= 7:
        return 'Rare'
    return 'Common'

# تعبئة البيانات
for idx, name in enumerate(results, start=2):
    base = ''.join(c for c in name.lower() if c.isalpha() or c in 'æøêïç')
    length = len(base)
    pattern = choose_pattern(length) if length in [3,4,5,6] else '-'
    score = score_name(base)
    has_symbol = any(s in name for s in symbols)
    rarity = rarity_from_score(score, has_symbol)

    ws.cell(row=idx, column=1, value=idx-1)
    ws.cell(row=idx, column=2, value=name)
    ws.cell(row=idx, column=3, value=length)
    ws.cell(row=idx, column=4, value=pattern)
    ws.cell(row=idx, column=5, value=score)
    ws.cell(row=idx, column=6, value=rarity)
    ws.cell(row=idx, column=7, value='Yes' if has_symbol else 'No')

    fill = PatternFill('solid', fgColor=rarity_colors[rarity])
    for col in range(1, 8):
        ws.cell(row=idx, column=col).fill = fill

# عرض الأعمدة
widths = {1:8, 2:22, 3:10, 4:12, 5:10, 6:15, 7:10}
for col, width in widths.items():
    ws.column_dimensions[get_column_letter(col)].width = width

# تجميد الصف الأول
ws.freeze_panes = 'A2'

# ورقة إحصائيات
stats = wb.create_sheet(title='Statistics')
stats['A1'] = 'Total Names'
stats['B1'] = len(results)

from collections import Counter
rarity_count = Counter()

for name in results:
    base = ''.join(c for c in name.lower() if c.isalpha() or c in 'æøêïç')
    score = score_name(base)
    has_symbol = any(s in name for s in symbols)
    rarity_count[rarity_from_score(score, has_symbol)] += 1

row = 3
for rarity in ['Common', 'Rare', 'Epic', 'Legendary']:
    stats.cell(row=row, column=1, value=rarity)
    stats.cell(row=row, column=2, value=rarity_count[rarity])
    row += 1

# حفظ الملف
wb.save('UNIQUEULTIMATEV366.xlsx')
print('Saved to generated_names.xlsx')